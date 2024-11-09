import openpyxl
import os

# Path to the test data file
test_data_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "TestData", "login_data.xlsx")


# Function to get the row count in a sheet
def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


# Function to get the column count in a sheet
def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


# Function to read data from a specific cell in a sheet
def read_data(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_num).value


# Function to write data to a specific cell in a sheet
def write_data(file, sheet_name, row_num, column_num, value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_num).value = value  # Writing the value
    workbook.save(file)  # Saving the workbook after writing data


# Test code to show the functions working
if __name__ == "__main__":
    # Example usage
    sheet_name = "Sheet1"

    print(f"Row count: {get_row_count(test_data_file, sheet_name)}")
    print(f"Column count: {get_column_count(test_data_file, sheet_name)}")

    # Reading data from row 2, column 1
    print(f"Data in cell (3,1): {read_data(test_data_file, sheet_name, 3, 1)}")

    # Writing new data to row 2, column 1
    write_data(test_data_file, sheet_name, 2, 1, "new_user@example.com")
    print("Updated cell (2,1) with new data.")

    # Verifying the updated data
    print(f"Data in updated cell (2,1): {read_data(test_data_file, sheet_name, 2, 1)}")
